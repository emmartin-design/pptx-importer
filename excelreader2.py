from openpyxl import load_workbook
import pandas as pd

# Local Modules
import variables as v


# defining variables
chart_types = v.chart_types
data_error = v.data_error
error_dict = v.error_dict
exception_list = v.exception_list


def longest_val(lst):
    longest, shortest = lst[0], lst[0]
    for val in lst:
        if len(val) > len(longest):
            longest = val
        elif len(val) < len(shortest):
            shortest = val
    return longest, shortest


def str_cleanup(string, upper=False):
    clean_str = str(string)
    for mark in ["'", ":"]:
        clean_str = clean_str.replace(mark, "", 10)
    if upper:
        clean_str = clean_str.upper().strip()
    return clean_str


def remove_parentheticals(og_dict):
    new_dict = og_dict.copy()
    for key in og_dict:
        for val_idx, val in enumerate(og_dict[key]):
            if type(val) == str:
                open_par_idx, close_par_idx = val.find('('), val.find(')')
                new_val = str_cleanup(val.replace(val[open_par_idx:(close_par_idx + 1)], ''))
                new_dict[key][val_idx] = new_val.strip()
    return new_dict


# Create new dataframe containing only the country for each report and the global average
def filter_countries(df, country, bases):
    base_idx = df.columns.get_loc(country)
    new_bases = bases[base_idx]
    country_df = df[['Global Average', country]].copy
    return country_df, new_bases


def remove_blank_values(original_dict):
    new_dict = {key: value for (key, value) in original_dict.items() if len(value) > 0}
    return new_dict


def remove_uppercase_letters(original_dict):
    new_dict = original_dict.copy()
    for key in original_dict:
        for val in original_dict[key]:
            try:
                if val.isupper() and val not in v.common_abbreviations:
                    new_dict[key].remove(val)
            except AttributeError:
                pass
    return new_dict


def configure_pages(page_data, tag=None):
    for page in page_data:
        chart_count, table_count = 0, 0
        for chart in page_data[page]:
            if chart != 'config':
                if page_data[page][chart]['config']['intended chart'] == 'TABLE':
                    table_count += 1
                else:
                    chart_count += 1
        page_data[page]['page config'] = v.assign_page_config(chart_count, table_count, tag=tag)
    return page_data


def generate_series_names(og_dict, config):
    new_dict = og_dict.copy()
    series_names = ['categories']
    for idx, key in enumerate(og_dict):
        if idx != 0:
            if all(type(value) == str for value in og_dict[key]):
                [series_names.append(name) for name in og_dict[key] if name not in series_names]
                config['vertical series'] = True
                new_dict.pop(key, None)
            else:
                if type(og_dict[key][0]) == str:
                    series_names.append(og_dict[key][0])
                    new_dict[key].pop(0)
                else:
                    if not config['vertical series']:  # Must check, else too many series names
                        series_names.append(('placeholder ' + str(idx)))
    return new_dict, config, series_names


def format_series(og_dict, series_names):
    new_dict = {}
    for key, name in zip (og_dict, series_names):
        new_dict[name] = og_dict[key]
    return new_dict


def format_vertical_series(og_dict, series_names):
    new_dict = {}
    for key_idx, key in enumerate(og_dict):
        if key_idx == 0:
            new_dict[series_names[0]] = og_dict[key]
            series_names.pop(0)
        else:
            for name_idx, name in enumerate(series_names):
                new_dict[name] = og_dict[key][name_idx::len(series_names)]
    return new_dict


def create_series_filter_list(df, config):
    cat_filter_list = []
    series_filter_list = []
    if 'Base' in df.columns:
        if df['Base'].sum() / len(df.index) == 1 or config['vertical series']:
            series_filter_list.append('Base')
        else:
            df.rename(columns={'Base': 'Overall'}, inplace=True)

    mean_list = []
    for i in df.index:
        if i in ['Mean', 'Median', 'Base']:
            mean_list.append(i)
    if len(df.index) > len(mean_list):
        cat_filter_list += mean_list
    else:
        cat_filter_list.append('Base')

    return cat_filter_list, series_filter_list


def frame_filter(og_df, category_filter, series_filter):
    og_series_names = og_df.columns.tolist()
    og_cat_names = og_df.index.tolist()

    for name_idx, name in enumerate(og_series_names):
        if name in series_filter:
            og_series_names.pop(name_idx)

    for name_idx, name in enumerate(og_cat_names):
        if name in category_filter:
            og_cat_names.pop(name_idx)

    new_df = og_df.loc[og_cat_names, og_series_names]

    return new_df


def convert_to_dataframe(og_dict, config, country):
    try:
        df = pd.DataFrame.from_dict(og_dict)
        df.set_index('categories', inplace=True)

        # Selects only specific country data for country reports
        if country is not None:
            df, config['bases'] = filter_countries(df.copy(), country, config['bases'].copy())

    except ValueError:
        # Data length mismatch causes errors
        error_data = {"ERROR": ['CHECK', 'DATA', 'SHEET'], "COLLECTION": [0, 0, 0], "CHECK DATA": [0, 0, 0]}
        df = pd.DataFrame.from_dict(error_data)
        df.set_index('ERROR', inplace=True)
        config['error list'].append('100')
        config['notes'].append(str(og_dict))

    return df, config


###############################################

def clean_up_data(data_dict, config, country):  # Checks data for gaps, formats based on config

    # Remove empty series from dictionary
    data_dict = remove_blank_values(data_dict)

    # Remove Uppercase letters from dictionary
    data_dict = remove_uppercase_letters(data_dict)

    # Detect Series Names, including vertical series
    data_dict, config, series_names = generate_series_names(data_dict, config)

    # Format series
    if config['vertical series']:
        data_dict = format_vertical_series(data_dict, series_names)
    else:
        data_dict = format_series(data_dict, series_names)

    # Clean up Category and Series names
    data_dict = remove_parentheticals(data_dict)

    # Convert to DataFrame
    if len(data_dict) > 0:
        df, config = convert_to_dataframe(data_dict, config, country)
    else:
        return None, config

    # Create list of filtered out series and categories (always includes Base)
    cat_filter_list, series_filter_list = create_series_filter_list(df, config)

    # Filter Series and Columns
    df = frame_filter(df, cat_filter_list, series_filter_list)

    return df, config


def data_collector(app, file, pptx_name, country=None):
    wb = load_workbook(filename=file, data_only=True)
    page_data = {}
    most_recent_tab_color = None
    page_counter = 0

    sheet_data = {}
    for sheet_idx, sheet in enumerate(wb.worksheets):
        msg_for_ui = 'Reading ' + str(wb.sheetnames[sheet_idx])
        v.log_entry(msg_for_ui, level='info', app_holder=app, fieldno=1)

        # Create default config data for chart and add it to the sheet data dictionary
        sheet_data[sheet_idx] = {}
        config = v.assign_chart_config(state=sheet.sheet_state, tab_color=sheet.sheet_properties.tabColor)
        config['notes'].append(sheet)

        # Create dictionary to collect data. Can only be one 'categories' per sheet
        frame_data = {}
        column_list = []  # Used for bases
        base_check = False  # If 'base' has already been found, ignore future 'base' selections

        # Iterate over excel by column to collect data
        for col_idx, col in enumerate(sheet.iter_cols()):
            series_lst = []  # Resets the list for new series
            cell_memory = None  # Recalls most recent cell

            # Check each cell for commands or color coding
            for cell_idx, cell in enumerate(col):
                cell_val, cell_color = sheet[cell.coordinate].value, cell.fill.start_color.index
                cell_format = sheet[cell.coordinate].number_format
                cell_val_cleaned_up = str_cleanup(cell_val, True)  # Removes extra spaces, capitalizes strings

                if cell_val_cleaned_up in config.keys():  # Checks for commands
                    config[cell_val_cleaned_up] = True

                elif cell_val_cleaned_up in chart_types:  # detects chart type
                    config['intended chart'], config['chart chosen'] = cell_val_cleaned_up, True

                elif cell_color == 4:  # detects data question for footer/chart title
                    config['title question'].append(cell_val)  # Update split to question and title

                elif cell_color == 5:  # detects bases for footer
                    column_list.append(col_idx)  # Tracks column indexes of bases to align them with correct axis
                    try:
                        config['bases'].append(('{:,}'.format(cell_val)))  # Adds commas to base numbers when possible
                    except ValueError:
                        config['bases'].append(cell_val)

                elif cell_color == 7:  # detects all data
                    if cell_val is not None:
                        if cell_format in ['0.0', '0%']:
                            try:
                                series_lst.append(float(cell_val))
                            except ValueError:
                                if cell_val_cleaned_up in ['*', '-']:  # Separate to catch formatting errors
                                    series_lst.append(float(0))
                        elif cell_format in ['@', 'General']:
                            try:  # Filters out Integers
                                int_test = int(cell_val)
                                if cell_format == '@':
                                    series_lst.append(str(cell_val))
                            except ValueError:
                                if cell_val_cleaned_up in exception_list:
                                    if cell_val == cell_memory:
                                        series_lst.append(float(0))
                                else:
                                    series_lst.append(str(cell_val))
                        if '*' in cell_val_cleaned_up:
                            config['directional check'] = True

                cell_memory = cell_val

            frame_data[col_idx] = series_lst  # Data is still pretty dirty at this point.

        # Cleans up collected data
        df, config = clean_up_data(frame_data, config, country)
        if df is None:
            continue

        # Applies transposition
        if config['*TRANSPOSE']:
            df_t = df.transpose()
            df = df_t  # replaces data with transposed version
            config['notes'].append("Chart Transposed")

        # If chart is Top Box, create sum column
        top_box_name = 'Top ' + str(len(df.columns)) + ' Box'  # Only used for Top Box charts
        if config['*TOP BOX']:
            if config['intended chart'] not in ['STACKED BAR', 'STACKED COLUMN']:
                config['intended chart'] = 'STACKED BAR'
            df[top_box_name] = df.sum(axis=1)

        # Sort data by first column in DF after transposing, or if Top Box, sum first, then first column
        sort_parameters = [config['*SORT'], config['*TOP 5'], config['*TOP 10'], config['*TOP 10']]
        top_ranges = {'*TOP 5': 5, '*TOP 10': 10, '*TOP 20': 20}

        if True in sort_parameters:
            config['notes'].append("Chart sorted by " + str(df.columns[0]))
            if config['*TOP BOX']:
                df_s = df.sort_values(by=[top_box_name, df.columns[0]], ascending=[False, False])
            else:
                df_s = df.sort_values(by=[df.columns[0]], ascending=[False])
            df = df_s

        # Truncates data to top selection.
        # If multiple top selections are added, should use the largest.
        for top_range in top_ranges:
            modifier = -1
            if config[top_range]:
                df_col_list = df[df.columns[0]].tolist()
                for value in df_col_list:
                    if value == df.iat[top_ranges[top_range], 0]:
                        modifier += 1
                truncated_df = df.iloc[0:(top_ranges[top_range] + modifier), :]
                df = truncated_df

        # Used to highlight highest values in table row after transposing
        config['max values'] = df.max(axis=1).values.tolist()

        # Used to create percents from whole numbers
        if config['*FORCE PERCENT']:
            config['percent check'] = True
            df = df.apply(lambda row: (row / 100), axis=1)

        # Check Tab color against most recent. If different, create new page.
        if config['tab color'] is None or config['tab color'] != most_recent_tab_color:
            page_counter += 1
            page_data[page_counter] = {}

        most_recent_tab_color = config['tab color']
        page_data[page_counter][sheet_idx] = {
            'frame': df,
            'config': config
        }

    # Count data visualizations and assign page config
    formatted_page_data = configure_pages(page_data, tag=pptx_name)

    return formatted_page_data
