from openpyxl import load_workbook
from collections import OrderedDict
import pandas as pd
import logging
from datetime import datetime

#Local Module
import variables as v

# defining variables
chart_types = v.chart_types
data_error = v.data_error
error_dict = v.error_dict
exception_list = v.exception_list

# pandas can open and return all worksheets as a dictionary.
# Use pandas to return data once the interface is built out.
def readsheet(file, wksht):
    df = pd.read_excel(file, sheet_name=wksht)
    return df


def strcleanup(string, upper=False):
    cleanstr = str(string).replace(':', '', 10)
    if upper:
        cleanstr = cleanstr.upper().strip()
    return cleanstr


def sheet_to_df(file):  # Used for position-based report types
    exceldata = pd.read_excel(file, index_col=0, sheet_name=0)
    df = pd.DataFrame(exceldata)
    textdf = pd.read_excel(file, sheet_name=1)
    textdata = textdf.iloc[:, 0].to_list()
    textdata.insert(0, textdf.columns[0])
    loggingmsg = 'Converted '  + file + ' to dataframe'
    logging.info(loggingmsg)
    return df, textdata


# Removes any blank entries from the data dictionary
def scrubber(dataog, app=None):
    data = dataog.copy()
    for slidecount, slide in enumerate(dataog):
        if app != None:
            msg_for_ui = 'Scrubbing Data — ' + str(trueround(((slidecount / len(dataog)) * 100), 0)) + '%'
            v.log_entry(msg_for_ui, app_holder=app, fieldno=1)
        page_config = dataog[slide]['page config']
        page_config['number of tables'] = 0
        page_config['number of charts'] = 0
        for card in dataog[slide]:
            if card != 'page config':
                config = dataog[slide][card]['config']
                if config['intended chart'] == 'TABLE':
                    page_config['number of tables'] += 1
                elif config['has data']:
                    page_config['number of charts'] += 1

        datavizcount = page_config['number of charts'] + page_config['number of tables']

        if page_config['page title'] != None:
            copycount = 1 + len(page_config['page copy'])
        else:
            copycount = len(page_config['page copy'])

        if datavizcount + copycount == 0:
            del data[slide]
    return data


def readbook(app, file, pptxname, country = None):
    wb = load_workbook(filename=file, data_only=True)
    slide_data = OrderedDict()

    combinecount = 1  # Indicates how many charts per slide
    most_recent_tabcolor = None

    slidecount = 0

    for sheetcount, wksht in enumerate(wb.worksheets):
        msg_for_ui = 'Reading ' + str(wb.sheetnames[sheetcount])
        v.log_entry(msg_for_ui, app_holder=app, fieldno=1)
        if wksht.sheet_state == "visible":
            tabcolor = wksht.sheet_properties.tabColor
        if sheetcount not in slide_data:
            slide_data[sheetcount] = {}  # sets up dictionary for future use
            slide_data[sheetcount]['page config'] = v.assign_page_config(tag=pptxname)

        # Creates new config dict for each tab
        config = v.assign_chart_config(state=wksht.sheet_state, tab_color=wksht.sheet_properties.tabColor)

        if wksht.sheet_state == "visible":
            tabcolor = wksht.sheet_properties.tabColor  # Used for combining charts

            config['notes'].append(wksht)
            framedata = OrderedDict()  # Used to create dataframe
            framedata['categories'] = []
            seriesnamelist = []
            colchecklst = []

            for col in wksht.iter_cols():  # Colcheckno is used to see if bases are vertical
                colcheckno = 0  # Resets to 0 each column. If higher than 0, indicates vertical positioning.

                serieslist = []  # blanks out series for every new column
                for cell in col:
                    cellval = wksht[cell.coordinate].value
                    cellcolor = cell.fill.start_color.index

                    # update config
                    cellval_c = strcleanup(cellval, True)  # Used for commands and chart types, not copy.
                    if cellval_c in config.keys():
                        config[cellval_c] = True
                    elif cell.font.underline == 'single':  # detects slide copy command
                        config['copy'].append(cellval)
                    elif cellval_c in chart_types:  # detects chart type command
                        config['intended chart'], config['chart chosen'] = cellval_c, True

                    # color based selections
                    elif cellcolor == 4:  # detects data question for footer/chart title
                        config['title question'].append(cellval)  # Update split to question and title

                    elif cellcolor == 5:  # detects bases for footer
                        colchecklst.append(colcheckno)  # Used to determine base labels (rows v. cols)
                        try:
                            config['bases'].append(('{:,}'.format(cellval)))  # Adds commas to base numbers
                        except ValueError:
                            config['bases'].append(cellval)
                        colcheckno += 1

                    elif cellcolor == 7:  # detects categories
                        if cellval is None:
                            config['error list'].append('106')
                            logging.warning('106 ERROR: blank category cell collected')
                        else:
                            cat = strcleanup(cellval)
                            framedata['categories'].append(cat)

                    elif cellcolor == 8:  # Collects series name (TO UPDATE: merge with series list with POP?)
                        if cellval is not None:
                            series_name = strcleanup(cellval)
                        else:
                            series_name = "ERROR_PLACEHOLDER"
                        seriesnamelist.append(series_name)

                    elif cellcolor == 9:  # detects data
                        config['has data'] = True
                        if config['*FORCE FLOAT'] or config['*FORCE CURRENCY']:
                            config['percent check'] = False
                        if cellval is None:  # Separate check for error reporting
                            serieslist.append(0)
                            config['error list'].append('106')
                            logging.warning('106 ERROR')

                        if config['*FORCE PERCENT'] is True:
                            if cellval in exception_list:
                                percentvalue = 0
                            else:
                                percentvalue = float(cellval) / 100
                            serieslist.append(percentvalue)
                        elif type(cellval) != float:
                            if type(cellval) is int and cellval > 1:
                                serieslist.append(float(cellval))
                            if cellval == 1 and config['percent check'] is True:
                                serieslist.append(1.0)
                            elif cellval == 0 or cellval_c in exception_list:
                                serieslist.append(0.0)
                            elif '=' in str(cellval):
                                config['error list'].append('107')  # TO UPDATE: Read formulas
                                logging.warning('107 ERROR')
                                serieslist.append(0.0)

                        else:
                            serieslist.append(cellval)

                if len(serieslist) > 0:  # Checks if series were collected
                    if len(framedata['categories']) != len(serieslist):
                        config['error list'].append('100')
                    framedata[series_name] = serieslist  # Adds series to data



            # Cleanup series names for vertical series (happens after initial collection)
            if len(seriesnamelist) > (len(framedata) - 1):
                newdata = OrderedDict()
                newdata['categories'] = framedata['categories'].copy()

                # The following appends non-repeating series names.
                # The set command doesn't maintain order, causing major issues.
                newseriesnames = []
                for name in seriesnamelist:
                    if name not in newseriesnames:
                        newseriesnames.append(name)
                lststep = len(newseriesnames)  # determines the step number for the data reorg

                # Find series name in original data for parsing
                if 1 < len(framedata) < 3:
                    for key in framedata:
                        if key != 'categories':
                            keyholder = key
                else:
                    logging.error('Major data issue on ' + str(wksht) + '. Check series selections.')

                for lstidx, series in enumerate(newseriesnames):
                    newdata[series] = []
                    for idx in range(lstidx, len(framedata[keyholder]), lststep):
                        newdata[series].append(framedata[keyholder][idx])

                framedata = newdata.copy()  # Replaces erroroneous data with reconfigured data
                if '100' in config['error list']:  # Removes previously applied error for vertical series charts
                    config['error list'].remove('100')

            # Creates a dataframe from the collected data. The dictionary is dropped from here on in.
            try:
                if country is not None:  # If this is a country report, it will filter out countries
                    newbases = []
                    df_og = pd.DataFrame.from_dict(framedata)
                    basedict = {'categories': 'bases'}
                    for idx, col in enumerate(df_og.columns, start=-1): # Create dictionary of bases
                        if col != 'categories':
                            basedict[col] = config['bases'][idx]
                    df_w_bases = df_og.append(basedict, ignore_index=True)
                    new_df = df_w_bases[['categories', 'Global Average', country]].copy()
                    baseholder = new_df.iloc[-1:].values.tolist()
                    for base in baseholder:
                        for b in base:
                            if b != 'bases':
                                newbases.append(b)
                    config['bases'] = newbases
                    new_df.drop(new_df.tail(1).index,inplace=True)
                    df = new_df
                    df.set_index('categories', inplace=True)
                else:
                    df = pd.DataFrame.from_dict(framedata)
                    df.set_index('categories', inplace=True)

                config['max values'] = df.max(axis=1).values.tolist()

            except:  # This drops in dummy data if the data doesn't fit into the dataframe.
                df = pd.DataFrame.from_dict(data_error)
                config['error list'].append('100')
                logging.warning('100 ERROR')
                df.set_index('categories', inplace=True)


            # Base Collection—Allows multiple bases
            data_base_list = []
            basecount = len(config['bases'])
            config['directional check'] = basecount > 0 and '35' in config['bases']
            colcheck = 1 in colchecklst  # True if vertical base selection
            basestr = ''  # Used for complex country report base formatting
            for nameidx, base in enumerate(config['bases']):
                if basecount == 1:
                    data_base_list.append('Base: ' + str(base))
                else:
                    if country is None:
                        if colcheck:  # Means bases refers to categories, not series
                            data_base_list.append(str(framedata['categories'][nameidx]) + ' Base: ' + str(base))
                        else:
                            data_base_list.append(str(df.columns[nameidx]) + ' Base: ' + str(base))

                    else:  # TO UPDATE: This will break is GA comes after country in data.
                        if df.columns[nameidx] == 'Global Average':
                            basestr = ('Base: ' + str(base) + '(global average)')
                        elif df.columns[nameidx] == country:
                            basestr = basestr + (' and ' + str(base) + '(' + str(df.columns[nameidx]) + ')')
                            data_base_list.append(basestr)

            config['bases'] = data_base_list

            if len(df.index) > 10 and config['intended chart'] == 'PIE':
                config['intended chart'] = 'BAR'

            #  Transposes Data
            if config['*TRANSPOSE'] is True:
                df_t = df.transpose()
                df = df_t  # replaces data with transposed version
                config['notes'].append("Chart Transposed")
                logging.info("Chart Transposed")

            # If chart is Top Box, create sum column
            if config['*TOP BOX'] == True:
                if config['intended chart'] not in ['STACKED BAR', 'STACKED COLUMN']:
                    config['intended chart'] = 'STACKED BAR'
                topboxname = 'Top ' + str(len(df.columns)) + ' Box'
                df[topboxname] = df.sum(axis=1)

            #  Sort data with Pandas. Does not work if category/series mismatch error thrown earlier
            sort_parameters = [config['*SORT'], config['*TOP 5'], config['*TOP 10'], config['*TOP 10']]
            top_ranges = {'*TOP 5': 5, '*TOP 10': 10, '*TOP 20': 20}

            if True in sort_parameters and '100' not in config['error list']:
                if len(framedata) > 0:
                    colname = df.columns[0]
                    s_note = "Chart sorted by " + str(colname)
                    config['notes'].append(s_note)
                    if config['*TOP BOX'] == True:
                        df_s = df.sort_values(by=[topboxname, colname], ascending=[False, False])
                    else:
                        df_s = df.sort_values(by=[colname], ascending=[False])
                    df = df_s

            # If top selections are true:
            for top_range in top_ranges:
                if config[top_range]:
                    truncated_df = df.iloc[0:top_ranges[top_range], : ]
                    df = truncated_df

            #  Combine Check looks at tab color to see if new tab combined with previous on slide
            if combinecount > 6:
                logging.warning('Too many charts for combining. Cannot combine more than 6.')
                combinecount = 1
            if tabcolor is not None:
                if tabcolor == most_recent_tabcolor:  # compares tab to most recent tab
                    combinecount += 1
                    slidecount -= 1  # Keeps slide count accurate for error reporting
                else:
                    combinecount = 1
                most_recent_tabcolor = tabcolor
            else:
                combinecount = 1

            slide_data[slidecount][combinecount] = {'config': config, 'frame': df}
            slidecount += 1

    # Scrub Blanks
    v.log_entry('Scrubbing Data', app_holder=app, fieldno=1)
    slide_data_sd = scrubber(slide_data, app)
    v.log_entry('Data Collection Complete', app_holder=app, fieldno=1)

    return slide_data_sd


def scorecard(file, config):
    # Copies config to allow changes
    report_config = config.copy()

    # Reads data from user-selected sheet and creates dataframe
    exceldata = pd.read_excel(file, sheet_name=0)
    df = pd.DataFrame(exceldata)

    # Reads text from excel sheets maintained by software admin
    explainer_text_1 = pd.read_excel(report_config['additional files'], sheet_name=0)

    try:  # Only used for LTO pages
        explainer_text_2 = pd.read_excel(report_config['additional files'], sheet_name=1)
        explainer_text_3 = pd.read_excel(report_config['additional files'], sheet_name=2)
    except IndexError:
        pass

    # Set up data
    df.columns = v.scorecard_dfcols[report_config['report type']]
    df.set_index('Concept', inplace=True)

    for col in df:
        for idx, val in enumerate(df[col]):
            if type(val) is not str:
                df[col][idx] = v.trueround(val, 2)

    # Add Placeholder data for columns not in spreadsheet
    for col in ['Designation', 'Dine-In', 'Takeout', 'Delivery']:
        df[col] = 0

    # Page counting and dictionary update functions
    def new_page(page_counter, slide_data, parameters=None):
        page_counter += 1
        slide_data[page_counter] = {}
        slide_data[page_counter]['page config'] = v.assign_page_config()
        if parameters is not None:
            for parameter in parameters:
                slide_data[page_counter]['page config'][parameter] = parameters[parameter]

        return page_counter, slide_data

    # Set up page dictionary
    slide_data = {}
    page_counter = 0  # Allows streamlining the content for two similar report structures.

    # Create Cover #########################################################################
    cover_parameters = {
        'function': 'cover',
        'page title': report_config['report title'],
        'page copy': report_config['table of contents']
    }
    page_counter, slide_data = new_page(page_counter, slide_data, cover_parameters)

    if report_config['report type'] == 'lto':
        # Create Methodology Page ###########################################################
        intro_title = explainer_text_1.columns[0]
        text = explainer_text_1[intro_title].values.tolist()
        text.insert(0, intro_title)

        methodology_parmeters = {
            'function': 'text',
            'section tag': 'Methodology and Definitions',
            'page copy': text
        }

        page_counter, slide_data = new_page(page_counter, slide_data, methodology_parmeters)


        # Create Definitions Page ###########################################################
        intro_title = explainer_text_2.columns[0]
        text = explainer_text_2[intro_title].values.tolist()
        text.insert(0, intro_title)

        definitions_parameters = {
            'function': 'text',
            'section tag': 'Methodology and Definitions',
            'page copy': text
        }

        page_counter, slide_data = new_page(page_counter, slide_data, definitions_parameters)

        # Create FAQ Page ###########################################################
        intro_title = explainer_text_3.columns[0]
        text = explainer_text_3[intro_title].values.tolist()
        text.insert(0, intro_title)

        faq_parameters = {
            'function': 'text',
            'section tag': 'Methodology and Definitions',
            'page copy': text
        }

        page_counter, slide_data = new_page(page_counter, slide_data, faq_parameters)

        # Create Intro Slide ##################################################################
        intro_parameters = {
            'function': 'intro',
            'section tag': report_config['report title'],
            'page copy': ['Recommended Action', 'Details and Qualifiers']
        }

        page_counter, slide_data = new_page(page_counter, slide_data, intro_parameters)


    elif report_config['report type'] == 'value':
        # Create Intro Slide ##################################################################
        intro_parameters = {
            'function': 'intro',
            'section tag': report_config['report title'],
            'page copy': ['Big Idea', 'Small Idea']
        }

        page_counter, slide_data = new_page(page_counter, slide_data, intro_parameters)

        # Create Explainer text page ##################################################################
        intro_title = explainer_text_1.columns[0]
        text = explainer_text_1[intro_title].values.tolist()
        text.insert(0, intro_title)

        explainer_parameters = {
            'function': 'text',
            'section tag': report_config['report title'],
            'page copy': text
        }

        page_counter, slide_data = new_page(page_counter, slide_data, explainer_parameters)

    # Create Benchmarking Table for both report types ##################################################################
    benchmarking_parameters = {
        'function': 'table',
        'section tag': 'Concept Scorecards and Benchmarking',
        'number of tables': 1,
        'full page chart': True
    }

    page_counter, slide_data = new_page(page_counter, slide_data, benchmarking_parameters)

    indexes_df = df[report_config['index column names']].copy()
    chart_title = 'CONCEPT BENCHMARKING'

    if report_config['report type'] == 'lto':
        for col in report_config['index column names']:
            indexes_df[col] = (indexes_df[col] * 100).astype(int)
        note = ['*Index score based on top-box response within daypart-mealpart']
    else:
        note = None

    idx_roundup_config = v.assign_chart_config(indexes=report_config['index ranges'], intended_chart='TABLE',
                                               has_data=True, banding='cols', note=note, force_int=True,
                                               title=chart_title)
    slide_data[page_counter][1] = {'frame': indexes_df, 'config': idx_roundup_config}

    # Create Scorecards
    for concept_idx, concept in enumerate(df.index):

        scorecard_parameters = {
            'section tag': 'Concept Scorecards and Benchmarking',
            'number of charts': report_config['scorecard chart count'],
            'page copy': [('/tag' + str(df.iloc[concept_idx,0]).upper()),
                          ('/h1' + str(df.index[concept_idx])), df.iloc[concept_idx,1]]
        }

        page_counter, slide_data = new_page(page_counter, slide_data, scorecard_parameters)

        # Set up charts
        for chart in range(1, (report_config['scorecard chart count'] + 1)):
            slide_data[page_counter][chart] = {}

        if report_config['report type'] == 'lto':
            note = ['/bDefinitions, 2nd Box and Top Box respectively:',
                    '/bPurchase Intent#: Likely or very likely to purchase',
                    '/bUniqueness#: Unique and very unique',
                    '/bDraw#: Likely or much more likely to order from',
                    '/bCraveability#: Think I would crave and would definitely crave']

            # PI, UNI, DRAW, CRV Stacked Bar
            pidf = df.loc[[concept], ['PI 2nd Box', 'PI Top Box', 'PI Top 2 Box']].copy()
            udf = df.loc[[concept], ['Uniqueness 2nd Box', 'Uniqueness Top Box', 'Uniqueness Top 2 Box']].copy()
            ddf = df.loc[[concept], ['Draw 2nd Box', 'Draw Top Box', 'Draw Top 2 Box']].copy()
            cdf = df.loc[[concept], ['CRV 2nd Box', 'CRV Top Box', 'CRV Top 2 Box']].copy()

            for f in [pidf, udf, ddf, cdf]:
                f.columns = ['2nd Box', 'Top Box', 'Top 2 Box']

            pudcdf = pd.concat([pidf, udf, ddf, cdf])
            pudcdf.index = ['Purchase Intent', 'Uniqueness', 'Draw', 'Craveability']
            slide_data[page_counter][1]['frame'] = pudcdf
            slide_data[page_counter][1]['config'] = v.assign_chart_config(intended_chart='STACKED COLUMN', top_box=True,
                                                                  legend_location='bottom', has_data=True, note=note)

            # Willingness to pay stat
            slide_data[page_counter][2]['frame'] = [str(f"${df.at[concept, 'Median willingness to pay']:.2f}"),
                                            'Median willingness to pay']
            slide_data[page_counter][2]['config'] = v.assign_chart_config(intended_chart='STAT', has_data=True)
            slide_data[page_counter]['page config']['has stat'] = True

            # Seasonality Pie
            chart_title = 'Seasonality (Would purchase ______ during the year)'
            sdf = df.loc[[concept], ['Anytime', 'Certain times', 'A few times']].copy()
            sdf_t = sdf.transpose()
            slide_data[page_counter][3]['frame'] = sdf_t
            slide_data[page_counter][3]['config'] = v.assign_chart_config(intended_chart='PIE', title=chart_title,
                                                                  has_data=True)
            # Repeat Trial Pie
            chart_title = 'Repeat trial'
            rtdf = df.loc[[concept], ['Once', 'Some Visits', 'Most Visits', 'Every Visit']].copy()
            rtdf_t = rtdf.transpose()
            slide_data[page_counter][4]['frame'] = rtdf_t
            slide_data[page_counter][4]['config'] = v.assign_chart_config(intended_chart='PIE', title=chart_title,
                                                                  has_data=True)

            # Percenticle Indexes
            index_ranges = {1: [1.22, 0.96], 2: [1.27, 1.00], 3: [1.16, 0.99], 4: [1.17, 0.98]}
            callout_sz = [0.5, 0.3]
            cols = ['Purchase Intent', 'Uniqueness', 'Draw', 'Craveability']
            left_vals = [4.7, 6.9, 9.06, 11.27]
            create_shape = [9.15, 0.3, 3.71, 1, 'white', 'INDEX', 'l', 'default']
            slide_data[page_counter]['page config']['callouts']['shape'] = create_shape
            for key_index, (left_val, col) in enumerate(zip(left_vals, cols), 1):
                if df.at[concept, col] > index_ranges[key_index][0]:
                    text_color = 'mint'
                elif df.at[concept, col] < index_ranges[key_index][1]:
                    text_color = 'mandarin'
                else:
                    text_color = 'default'
                callout = callout_sz + [left_val, 1, None, str(indexes_df.at[concept, col]), 'c', text_color]
                slide_data[page_counter]['page config']['callouts'][str(cols.index(col))] = callout

        elif report_config['report type'] == 'value':
            # Stack Bar Charts
            stacked_charts = {
                'Value': ['Val 2nd Box', 'Val Top Box', 'Val Top 2 Box'],
                'Purchase Intent': ['PI 2nd Box', 'PI Top Box', 'PI Top 2 Box'],
            }
            for stacked_idx, chart in enumerate(stacked_charts, start=1):
                stacked_df = df.loc[[concept], stacked_charts[chart]].copy()
                stacked_df.columns = ['2nd Box', 'Top Box', 'Top 2 Box']
                stacked_df.index = [chart]
                if stacked_idx == 1:
                    pref_color = 'blueberry'
                else:
                    pref_color = 'mint'
                slide_data[page_counter][stacked_idx]['frame'] = stacked_df
                slide_data[page_counter][stacked_idx]['config'] = v.assign_chart_config(intended_chart='STACKED COLUMN',
                                                                                top_box=True,
                                                                                preferred_color=pref_color,
                                                                                has_data=True,
                                                                                legend_location='bottom',
                                                                                note=v.scorecard_footercopy[report_config['report type']])

            # Draw, Craveability, Quality stacked bar
            ddf = df.loc[[concept], ['Draw 2nd Box', 'Draw Top Box', 'Draw Top 2 Box']].copy()
            cdf = df.loc[[concept], ['CRV 2nd Box', 'CRV Top Box', 'CRV Top 2 Box']].copy()
            qdf = df.loc[[concept], ['Quality 2nd Box', 'Quality Top Box', 'Quality Top 2 Box']].copy()
            for f in [ddf, cdf, qdf]:
                f.columns = ['2nd Box', 'Top Box', 'Top 2 Box']
            dcqdf = pd.concat([ddf, cdf, qdf])
            dcqdf.index = ['Draw', 'Craveability', 'Quality']

            slide_data[page_counter][3]['frame'] = dcqdf
            slide_data[page_counter][3]['config'] = v.assign_chart_config(intended_chart='STACKED COLUMN', top_box=True,
                                                                  preferred_color='mint', has_data=True,
                                                                  legend_location='bottom')

            # Pie Charts
            pie_charts = {
                'Good or very good value reasoning': ['Portion', 'Taste', 'Quality', 'Fits Budget'],
                'Order behavior': ['Only the offer', 'Additional foods and/or beverages', 'Unsure'],
                'Repeat trial': ['Once', 'Some visits', 'Most visits', 'Every visit']
            }

            for pie_idx, pie in enumerate(pie_charts, start=4):
                chart_title = pie
                pie_df = df.loc[[concept], pie_charts[pie]].copy()
                if pie_idx == 4:
                    pie_df['Other'] = 1 - pie_df.sum(axis=1)
                    pie_df.index = ['Base']
                    pref_color = 'blueberry'
                else:
                    pref_color = 'mint'
                pie_df_t = pie_df.transpose()
                slide_data[page_counter][pie_idx]['frame'] = pie_df_t
                tempconfig = v.assign_chart_config(intended_chart='PIE', preferred_color=pref_color,
                                                   title=chart_title, has_data=True)
                slide_data[page_counter][pie_idx]['config'] = tempconfig

            # Percenticle Indexes
            callout_sz = [0.42, 0.3]
            cols = ['Value Percentile', 'Purchase Intent Percentile', 'Draw Percentile',
                    'Craveability Percentile', 'Quality Percentile']
            left_vals = [4.85, 8.02, 10.45, 11.33, 12.07]
            slide_data[page_counter]['page config']['callouts']['shape'] = [9.15, 0.3, 3.71, 1, 'white', 'PERCENTILES', 'l', 'default']
            for left_val, col in zip(left_vals, cols):
                slide_data[page_counter]['page config']['callouts'][str(cols.index(col))] = callout_sz + [left_val, 1, None,
                                                                              str(df.at[concept, col]),
                                                                              'c', 'default']

    # Create Brand Fit Placeholder for LTOs
    brand_fit_parameters = {
        'function': 'full page chart',
        'section tag': 'Concept Scorecards and Benchmarking',
        'number of charts': 1,
        'callouts': {
            'shape': [10.5, 0.3, 0.5, 6.36, 'white', 'INDEX', 'l', 'default'],
            1: [0.42, 0.3, 2.2, 6.36, None, str(90), 'c', 'default']
        }
    }

    page_counter, slide_data = new_page(page_counter, slide_data, brand_fit_parameters)
    mock_df = pd.DataFrame.from_dict(v.mock_chart)
    mock_df.set_index('categories', inplace=True)
    slide_data[page_counter][1] = {}
    slide_data[page_counter][1]['frame'] = mock_df
    slide_data[page_counter][1]['config'] = v.assign_chart_config(intended_chart='STACKED COLUMN',
                                                                  title='BRAND X BRAND FIT BY CONCEPT',
                                                                  has_data=True, top_box=True,
                                                                  note=['Index based on top-box response'])

    # Create Off-Premise Potential Placeholder for LTOs
    if report_config['report type'] == 'lto':
        off_premise_parameters = {
            'function': 'table',
            'section tag': 'Concept Scorecards and Benchmarking',
            'number of tables': 1,
            'full page chart': True
        }

        page_counter, slide_data = new_page(page_counter, slide_data, off_premise_parameters)

        off_premise_df = df[report_config['off-premise potential names']].copy()
        chart_title = 'OFF-PREMISE POTENTIAL'
        note = ['Q: How would you eat this item? Select all that apply.',
                'Base: potential purchasers (top 2 box purchase intent)']

        off_premise_config = v.assign_chart_config(intended_chart='TABLE', has_data=True, banding='rows', note=note,
                                                   title=chart_title)
        slide_data[page_counter][1] = {'frame': off_premise_df, 'config': off_premise_config}

    else:
        pass

    # Create Demographics Table(s)
    demogsdf = df[['Male', 'Female', 'Gen Z', 'Millennials', 'Gen X', 'Baby Boomers',
                   '<$45K', '$45K-$99K', '$100K+', 'Black/African American',
                   'White', 'Hispanic/Latino']].copy()

    demogsdf_t = demogsdf.transpose()
    concepts = list(demogsdf_t.columns)
    chart_title = 'DEMOGRAPHICS OF POTENTIAL PURCHASERS'
    note = ['Potential purchasers=top 2 box purchase intent']

    for idx, concept in enumerate(concepts):
        concepts[idx] = '/h2' + str(idx + 1) + '. ' + concept

    demogsdf_t.columns = list(range(1, len(concepts) + 1))
    maxvals = demogsdf_t.max(axis=1).values.tolist()

    df_lst, concept_lst = [], []
    col_len = len(demogsdf_t.columns)
    if 21 > col_len > 10:
        divide_col_len = int((col_len / 2))
    elif col_len > 20:
        divide_col_len = int((col_len / 3))
    else:
        divide_col_len = col_len

    for col in range(0, col_len, divide_col_len):
        df_lst.append(demogsdf_t.iloc[:, col:col + divide_col_len])
        concept_lst.append(concepts[col:col + divide_col_len])

    for frame_idx, frame in enumerate(df_lst):
        page_counter, slide_data = new_page(page_counter, slide_data)
        slide_data[page_counter]['page config'] = v.assign_page_config(function='table and text')
        slide_data[page_counter]['page config']['section tag'] = 'Concept Scorecards and Benchmarking'
        slide_data[page_counter]['page config']['number of tables'] += 1
        slide_data[page_counter]['page config']['page copy'] = concept_lst[frame_idx]
        slide_data[page_counter][1] = {
            'frame': frame,
            'config': v.assign_chart_config(max_values=maxvals, intended_chart='TABLE', has_data=True,
                                            banding='cols', highlight=True, title=chart_title, note=note)
        }

    # Create last Page
    page_counter, slide_data = new_page(page_counter, slide_data)
    slide_data[page_counter]['page config'] = v.assign_page_config(function='end wrapper')
    slide_data[page_counter]['page config']['page copy'] = ["So. What's Next?",
                                                            "Need some more LTO guidance? Reach out to our experts."]
    slide_data[page_counter]['page config']['section tag'] = 'About Technomic'

    for chart in [1, 2, 3, 4]:
        slide_data[page_counter][chart] = {}
        if chart < 3:
            ee = 'lh'
        else:
            ee = 'jc'
        if (chart % 2) == 0:
            slide_data[page_counter][chart]['frame'] = v.employees[ee]
            slide_data[page_counter][chart]['config'] = v.assign_chart_config(intended_chart='STAT')
        else:
            slide_data[page_counter][chart]['frame'] = 'templates/import_resources/headshots/' + ee + '.jpg'
            slide_data[page_counter][chart]['config'] = v.assign_chart_config(intended_chart='PICTURE')


    slide_data_sd = scrubber(slide_data)

    return slide_data_sd


def dtv_reader(file, entertainment=None):
    exceldata = pd.read_excel(file, sheet_name=0)
    df = pd.DataFrame(exceldata)
    slide_data = OrderedDict()
    df.columns = v.scorecard_dfcols['dtv']
    df.set_index(['Account', 'Competitor No.'], inplace=True)

    month_variant = 2  # May need to change based on report data date

    title_months = v.month_lst(3, ((datetime.now().month) - month_variant))
    reportdate = str(title_months[-1]) + ' ' + str(datetime.now().year)

    # If additional values are provided, a third layout per account will be created
    layout_list = [1, 2]
    if entertainment is not None:
        layout_list.append(3)

    page = 1
    for account, account_df in df.groupby(level=0):
        for layout in layout_list:
            slide_data[page] = {}
            slide_data[page]['page config'] = v.assign_page_config()
            page_config = slide_data[page]['page config']
            page_config['section tag'] = str(account) + ' | Consumer Visit Tracker & Ignite Consumer | ' + reportdate
            if layout == 1:  # First page of account
                for chart in range(1, 6):
                    slide_data[page][chart] = {}
                page_config['number of charts'] = 3
                page_config['number of tables'] = 2
                page_config['function'] = 'dtv'

                # Sales Market Share Stacked Bar
                sms_df = account_df.loc[:, ['Competitor Brand', 'Sales Market Share']].copy()
                sms_df.set_index(['Competitor Brand'], inplace=True)
                sms_df_t = sms_df.transpose()

                # Set up values to replace data labels
                growthdict = {1:[], 2:[], 3:[], 4:[]}
                curr_df=account_df.loc[:, ['Sales Market Share']].copy()
                currtxt = curr_df['Sales Market Share'].values.tolist()
                smsg_df = account_df.loc[:, ['Chg Sales Market Share']].copy()
                growthtxt = smsg_df['Chg Sales Market Share'].values.tolist()
                for idx, item in enumerate(growthtxt):
                    currval = round((currtxt[idx] * 100), 1)
                    f = '(%s)' % round(growthtxt[idx], 1)
                    growthdict[idx] = [(str(currval) + '% ' + f)]

                # Create chart title
                share_title = 'SALES MARKET SHARE\n'
                for month in title_months:
                    if month != title_months[-1]:
                        share_title += (str(month) + '-')
                share_title += (' ' + str(datetime.now().year) + ' VS. ' + str(datetime.now().year - 1))

                note_txt = [
                    '* Percent shown reflect results from September 2020 YTD',
                    'Arrows indicate change from August YTD 2020 to September 2020 YTD'
                ]

                slide_data[page][1]['frame'] = sms_df_t
                slide_data[page][1]['config'] = v.assign_chart_config(intended_chart='100% STACKED BAR',
                                                                      legend_location='bottom', title=share_title,
                                                                      category_axis=False, has_data=True,
                                                                      pct_dec_places=1, label_txt=growthdict,
                                                                      note=note_txt)

                # YOY Tables
                table_titles = {
                    'sales': ['Competitor Brand', 'YOY Sales Qtr', 'YOY Sales Annual'],
                    'traffic': ['Competitor Brand', 'YOY Traffic Qtr', 'YOY Traffic Annual']
                }

                for t_idx, table in enumerate(table_titles, 2):
                    table_df = account_df.loc[:, table_titles[table]].copy()
                    table_df.set_index(['Competitor Brand'], inplace=True)
                    slide_data[page][t_idx]['frame'] = table_df
                    slide_data[page][t_idx]['config'] = v.assign_chart_config(intended_chart='TABLE', has_data=True, growth=True,
                                                                              pct_dec_places=1)

                # Column Charts
                column_charts = {
                    'KEY PERFORMANCE INDICATORS*': {
                        'cols': ['Competitor Brand', 'Past Week Trial', 'Redeemed Coupon', 'Ordered LTO',
                                 'Lapsed User'],
                        'pp_cols': ['Competitor Brand', 'Chg Past Week Trial', 'Chg Redeemed Coupon', 'Chg Ordered LTO',
                                    'Chg Lapsed User'],
                        'color': 'mint'

                    },
                    'ORDERING METHOD*': {
                        'cols': ['Competitor Brand', 'Order In Store', 'Order Drive thru', 'Order Pickup/Delivery'],
                        'pp_cols': ['Competitor Brand', 'Chg Order In Store', 'Chg Order Drive thru',
                                    'Chg Order Pickup/Delivery'],
                        'color': 'grape'
                    }
                }

                for ch_idx, chart in enumerate(column_charts, 4):
                    ch_df = account_df.loc[:, column_charts[chart]['cols']].copy()
                    ch_pp_df = account_df.loc[:, column_charts[chart]['pp_cols']].copy()
                    for frame in [ch_df, ch_pp_df]:
                        frame.set_index(['Competitor Brand'], inplace=True)

                    ch_lst = ch_df.transpose().values.tolist()
                    ch_pp_lst = ch_pp_df.transpose().values.tolist()
                    ch_dict = {}
                    for s_idx, series in enumerate(ch_pp_lst):
                        ch_dict[s_idx] = []
                        for idx, item in enumerate(ch_pp_lst[s_idx]):
                            ch_val = int((ch_lst[s_idx][idx] * 100))
                            if ch_pp_lst[s_idx][idx] > 0:
                                growthval = u"\u2191"
                            elif ch_pp_lst[s_idx][idx] < 0:
                                growthval = u"\u2193"
                            else:
                                growthval = ''
                            ch_dict[s_idx].append((growthval + u"\u000A" + str(ch_val) + '%'))

                    slide_data[page][ch_idx]['frame'] = ch_df
                    slide_data[page][ch_idx]['config'] = v.assign_chart_config(intended_chart='COLUMN',
                                                                               legend_location='bottom',
                                                                               title=chart, has_data=True,
                                                                               label_txt=ch_dict,
                                                                               preferred_color=column_charts[chart]['color'])

                # Chart Titles as callouts
                # Width, Height, Left, Top,
                callout_sz = [2.71, .22]
                cols = ['YEAR-OVER-YEAR SALES', 'YEAR-OVER-YEAR TRAFFIC']
                left_vals = [3.62, 8.44]
                for left_val, col in zip(left_vals, cols):
                    page_config['callouts'][str(cols.index(col))] = callout_sz + [left_val, 1.19, None, col, 'l', 'default']

            elif layout == 2: # Line Graph Page
                page_config['number of charts'] = 2
                page_config['function'] = 'two charts, no text'

                line_graphs = {
                    'sales': {
                        'cols':['Competitor Brand', 'SM1','SM2','SM3','SM4','SM5','SM6','SM7'],
                        'title': 'SALES PERFORMANCE\nROLLING THREE-MONTH SYSTEMWIDE'
                    },
                    'traffic': {
                        'cols': ['Competitor Brand', 'TM1','TM2','TM3','TM4','TM5','TM6','TM7'],
                        'title': 'TRAFFIC PERFORMANCE\nROLLING THREE-MONTH SYSTEMWIDE'
                    }
                }

                for g_idx, graph in enumerate(line_graphs, 1):
                    slide_data[page][g_idx] = {}
                    g_df = account_df.loc[:, line_graphs[graph]['cols']].copy()
                    g_df.columns = ['Competitor Brand'] + v.month_lst(7, ((datetime.now().month) - month_variant), (datetime.now().year))
                    g_df.set_index(['Competitor Brand'], inplace=True)
                    g_df_t = g_df.transpose()
                    slide_data[page][g_idx]['frame'] = g_df_t
                    slide_data[page][g_idx]['config'] = v.assign_chart_config(intended_chart='LINE', legend_location=None,
                                                                              title=line_graphs[graph]['title'], has_data=True,
                                                                              data_labels=False, pct_dec_places=1)

            elif layout == 3:  # Third page of account only appears if values provided in app
                page_config['number of charts'] = 0
                page_config['number of tables'] = 1
                page_config['function'] = 'table and text'

                # Provides Entertainment Body Text (Segment level only)
                entertainment_values = {
                    'segment': ['Quick Service', 'Fast Casual', 'Midscale', 'Casual Dining'],
                    'Provides TV/Entertainment': entertainment
                }

                bodycopy = ['/b/h2Provides Video/TV Entertainment', '/h2Top Box', '']
                tv_df = pd.DataFrame.from_dict(entertainment_values)
                tv_df.set_index(['segment'], inplace=True)

                for idx, competitor in enumerate(tv_df.index):
                    val_copy = str(tv_df.iloc[idx, 0])
                    bodycopy.append(('/h1' + val_copy + '%'))
                    bodycopy.append(('/h2' + competitor))
                    bodycopy.append((''))
                page_config['page copy'] = bodycopy

                # Consumer Metrics Table
                slide_data[page][1] = {}
                cmt = { 'col_lst': ['Competitor Brand', 'Overall Ambience/Atmosphere', 'Video/TV Entertainment',
                                  'This was the right place for the occasion',
                                  'Appropriateness for the variety of occasions'],
                        'chg_col_lst': ['Competitor Brand', 'Chg Overall Ambience/Atmosphere',
                                      'Chg Video/TV Entertainment', 'Chg This was the right place for the occasion',
                                      'Chg Appropriateness for the variety of occasions']
                }

                note_txt = [
                    '/mintGreen# = increased from year-end Q2 2020 to year-end Q3 2020',
                    '/mandarinOrange# = decreased from year-end Q2 2020 to year-end Q3 2020'
                ]

                cm_df = account_df.loc[:, cmt['col_lst']].copy()
                chg_cm_df = account_df.loc[:, cmt['chg_col_lst']].copy()
                for frame in [cm_df, chg_cm_df]:
                    frame.set_index(['Competitor Brand'], inplace=True)
                chg_cm_df_t = chg_cm_df.transpose()
                cm_df_t = cm_df.transpose()
                slide_data[page][1]['frame'] = cm_df_t
                slide_data[page][1]['config'] = v.assign_chart_config(intended_chart='TABLE', has_data=True,
                                                                      indexes=chg_cm_df_t,
                                                                      pct_dec_places=1, note=note_txt)

            page += 1

    slide_data_sd = scrubber(slide_data)
    return slide_data_sd



